import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter, cell


class CreateTemplate: 

    def createTemplate(self, file_path):

        workbook = openpyxl.load_workbook(file_path)
        anforderungen_sheet = workbook['Anforderungen']
        tools_sheet = workbook['Tools']
        übersicht_sheet = workbook.create_sheet('Übersicht')
        auswertungen_sheet = workbook.create_sheet('Auswertungen')
        anforderungen = []
        tools = []
        central_alignment = Alignment(wrap_text=True,horizontal='center', vertical='center')
        alphabet = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

        # Bedingte Formatierung
        green_color = DifferentialStyle(fill=PatternFill(bgColor='C6EFCE')) #green
        yellow_color = DifferentialStyle(fill=PatternFill(bgColor='FFEB9C')) #yellow
        red_color = DifferentialStyle(fill=PatternFill(bgColor='FFC7CE')) #red
        green = Rule(type="cellIs", dxf=green_color, formula = ['"Ja"'], operator='equal')
        yellow = Rule(type="cellIs", dxf=yellow_color, formula = ['"Vielleicht"'], operator='equal')
        red = Rule(type="cellIs", dxf=red_color, formula = ['"Nein"'], operator='equal')


        # Alle Anforderungen in einer Liste speichern
        for row in anforderungen_sheet.iter_rows(min_row=2, max_col=3, values_only=True):
            if row[0] is not None and row[2] == "Ja":
                anforderungen.append(row[0])

        # Alle Tools in einer Liste speichern 
        for row in tools_sheet.iter_rows(min_row=2, max_col=3, values_only=True):
            if row[0] is not None and row[2] == "Ja":
                tools.append(row[0])




        # Übersicht Anforderungen einfügen
        übersicht_sheet.row_dimensions[1].height = 100
        übersicht_sheet.cell(row=1, column=1).value = 'Tools \ Anforderungen'
        übersicht_sheet.cell(row=1, column=1).alignment =  Alignment(wrap_text=True,horizontal='center', vertical='center')
        übersicht_sheet.column_dimensions['A'].width = 25
        for i, anforderung in enumerate(anforderungen):
            if anforderung is not None:
                übersicht_sheet.cell(row=1, column=i+2).value = anforderung
                #  Tabellen-Headings Text horizontal und vertikal zentrieren, wrap-text und gegen den Uhrzeigersinn ausrichten
                übersicht_sheet.cell(row=1, column=i+2).alignment = Alignment(wrap_text=True,horizontal='center', vertical='center', textRotation=45)

        # Übersicht Tools einfügen
        for i, tool in enumerate(tools):
            if tool is not None:
                übersicht_sheet.cell(row=i+2, column=1).value = tool
                # Tool soll auf Tool-Box in Auswertungen verweisen
                ziel_zelle_row = 2 + tools.index(tool)*7
                übersicht_sheet.cell(row=i+2, column=1).hyperlink = f'#Auswertungen!B{ziel_zelle_row}'

        # Tabelle befüllen
        for j, tool in enumerate(tools):
            for i, anforderung in enumerate(anforderungen):
                verweis_column = alphabet[(1+i)]
                verweis_row = 4+j*7
                # Falls Inhalt leer, soll statt 0 ein leerer String angezeigt werden
                übersicht_sheet.cell(row=2+j, column=2+i).value = f"=IF(Auswertungen!{verweis_column}{verweis_row}=\"\",\"\",Auswertungen!{verweis_column}{verweis_row})"
                übersicht_sheet.cell(row=2+j, column=2+i).alignment = Alignment(horizontal='center')


        # Ergebnis-Zeile hinzufügen, die die Anzahl der Tools in der Tabelle ausgibt
        number_of_tools = len(tools)
        ergebnis_zeile = number_of_tools +2
        for i in range(len(anforderungen)+1):
            if i == 0:
                # übersicht_sheet.cell(column=1, row=ergebnis_zeile).value = len(tools)
                übersicht_sheet.cell(column=1, row=ergebnis_zeile).value = f'=SUBTOTAL(103,A2:A{number_of_tools+1})'
                übersicht_sheet.cell(column=1, row=ergebnis_zeile).font = Font(bold=True, color='FFFFFF')
            übersicht_sheet.cell(column=1+i, row=ergebnis_zeile).fill = openpyxl.styles.PatternFill(start_color='4F71BE', end_color='4F71BE', fill_type='solid')
            
        # Anzahl-Spalte hinzufügen, die anzeigt, wie viele Anforderungen ein Tool erfüllt
        total_spalte = len(anforderungen) + 2
        übersicht_sheet.cell(row=1, column=total_spalte).value ='Anzahl'
        übersicht_sheet.cell(row=1, column=total_spalte).alignment = Alignment(wrap_text=True,horizontal='center', vertical='center', textRotation=45)

        for i in range(len(tools)):
            übersicht_sheet.cell(column=total_spalte, row=2+i).fill = openpyxl.styles.PatternFill(start_color='4F71BE', end_color='4F71BE', fill_type='solid')
            übersicht_sheet.cell(column=total_spalte, row=2+i).font = Font(bold=True, color='FFFFFF')
            end_column = alphabet[len(anforderungen)]
            übersicht_sheet.cell(column=total_spalte, row=2+i).value = f'=COUNTIF(B{2+i}:{end_column}{2+i}, \"Ja\")'
            
        # Inhalt bedingte Formatierung
        # Bedingte Formatierung für Dropdown
        begin_bf_übersicht = auswertungen_sheet.cell(row=2, column=2)
        end_bf_übersicht = auswertungen_sheet.cell(row=2+len(tools), column=2+len(anforderungen))
        
        übersicht_sheet.conditional_formatting.add(f'{get_column_letter(begin_bf_übersicht.column)}{begin_bf_übersicht.row}:{get_column_letter(end_bf_übersicht.column)}{end_bf_übersicht.row}', green)
        übersicht_sheet.conditional_formatting.add(f'{get_column_letter(begin_bf_übersicht.column)}{begin_bf_übersicht.row}:{get_column_letter(end_bf_übersicht.column)}{end_bf_übersicht.row}', yellow)
        übersicht_sheet.conditional_formatting.add(f'{get_column_letter(begin_bf_übersicht.column)}{begin_bf_übersicht.row}:{get_column_letter(end_bf_übersicht.column)}{end_bf_übersicht.row}', red)




        rows = len(tools) +1
        columns = alphabet[len(anforderungen)+1] 
        tab = Table(displayName="übersicht", ref=f"A1:{columns}{rows}")
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                    showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        übersicht_sheet.add_table(tab)




        # Auswertungen

        # Spaltenbreite
        last_col = alphabet[(len(anforderungen)+1)]
        for i in range(len(anforderungen)+1):
            auswertungen_sheet.column_dimensions[alphabet[i]].width = 20

        # Zellen-Höhe
        begin = 2
        end = len(tools)
        zeile = begin
        for begin in range(end):
            for i in range(4):
                auswertungen_sheet.row_dimensions[zeile].height = 45
                zeile +=1
            zeile += 3



        # Schleife für jedes Tool
        for i in range(len(tools)):
            # Border erstellen
            for a in range(len(anforderungen)):
                if a == 0:
                    auswertungen_sheet.cell(row=i*7+2, column=2+a).border = Border(left=Side(style='thin'), top=Side(style='thin')) 
                    auswertungen_sheet.cell(row=i*7+5, column=2+a).border = Border(left=Side(style='thin'), bottom=Side(style='thin')) 
                    auswertungen_sheet.cell(row=i*7+2, column=2+a).alignment = central_alignment
                    for b in range(2):
                        auswertungen_sheet.cell(row=i*7+3+b, column=2+a).border = Border(left=Side(style='thin')) 
                elif a == (len(anforderungen)-1):
                    auswertungen_sheet.cell(row=i*7+2, column=2+a).fill = openpyxl.styles.PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')
                    auswertungen_sheet.cell(row=i*7+2, column=2+a).border = Border(right=Side(style='thin'), top=Side(style='thin')) 
                    auswertungen_sheet.cell(row=i*7+5, column=2+a).border = Border(right=Side(style='thin'), bottom=Side(style='thin')) 
                    for b in range(2):
                        auswertungen_sheet.cell(row=i*7+3+b, column=2+a).border = Border(right=Side(style='thin'))
                else:
                    auswertungen_sheet.cell(row=i*7+2, column=2+a).fill = openpyxl.styles.PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')
                    auswertungen_sheet.cell(row=i*7+2, column=2+a).border = Border(top=Side(style='thin')) 
                    auswertungen_sheet.cell(row=i*7+5, column=2+a).border = Border(bottom=Side(style='thin')) 
                    auswertungen_sheet.cell(row=i*7+2, column=2+a).alignment = central_alignment
                    auswertungen_sheet.cell(row=i*7+5, column=2+a).alignment = central_alignment
            # Tool in die Zelle schreiben und Hintergrundfarbe einstellen
            auswertungen_sheet.cell(row=i*7+2, column=2).value = tools[i]
            auswertungen_sheet.merge_cells(start_row=i*7+2, start_column=2, end_row=i*7+2, end_column=3)
            auswertungen_sheet.cell(row=i*7+2, column=2).fill = openpyxl.styles.PatternFill(start_color='305496', end_color='305496', fill_type='solid')
            auswertungen_sheet.cell(row=i*7+2, column=3).fill = openpyxl.styles.PatternFill(start_color='305496', end_color='305496', fill_type='solid')
            auswertungen_sheet.cell(row=i*7+2, column=2).font = Font(size=16, color='FFFFFF')
            
            # Anforderungen in die Zellen darunter schreiben
            for j in range(len(anforderungen)):
                auswertungen_sheet.cell(row=i*7+3, column=2+j).value = anforderungen[j]
                auswertungen_sheet.cell(row=i*7+3, column=2+j).alignment = central_alignment
                auswertungen_sheet.cell(row=i*7+4, column=2+j).alignment = central_alignment
                auswertungen_sheet.cell(row=i*7+3, column=2+j).font = Font(size=14, color='FFFFFF')
                auswertungen_sheet.cell(row=i*7+4, column=2+j).font = Font(size=14)
                auswertungen_sheet.cell(row=i*7+3, column=2+j).fill = openpyxl.styles.PatternFill(start_color='8EA9DB', end_color='8EA9DB', fill_type='solid')

                # Dropdown in jeder Zelle unter den Anforderungen erstellen
                cell = auswertungen_sheet.cell(row=i*7+4, column=2+j)
                # Dropdown Liste
                dv = DataValidation(type="list", formula1='"Ja,Vielleicht,Nein"', allow_blank=True)
                auswertungen_sheet.add_data_validation(dv)
                dv.add(cell)
                
            # Bedingte Formatierung für Dropdown
            begin_bf = auswertungen_sheet.cell(row=i*7+4, column=2)
            end_bf = auswertungen_sheet.cell(row=i*7+4, column=(1+ len(anforderungen)))
        
            auswertungen_sheet.conditional_formatting.add(f'{get_column_letter(begin_bf.column)}{begin_bf.row}:{get_column_letter(end_bf.column)}{end_bf.row}', green)
            auswertungen_sheet.conditional_formatting.add(f'{get_column_letter(begin_bf.column)}{begin_bf.row}:{get_column_letter(end_bf.column)}{end_bf.row}', yellow)
            auswertungen_sheet.conditional_formatting.add(f'{get_column_letter(begin_bf.column)}{begin_bf.row}:{get_column_letter(end_bf.column)}{end_bf.row}', red)

        new_path = file_path.replace("Template", "IndivTemplate")
        print(new_path)
        workbook.save(new_path)